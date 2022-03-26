from openpyxl import load_workbook
from selenium import webdriver
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import msvcrt
import threading

pinFlag = 0
loginFlag = 0


class detectThread(threading.Thread):
    def __init__(self, threadID, name, xpath, driver, timeout, flag):
        threading.Thread.__init__(self)
        # self.result = False
        self.threadID = threadID
        self.name = name
        self.xpath = xpath
        self.driver = driver
        self.flag = flag
        self.timeout = timeout

    def run(self):
        print("开始线程：" + self.name)
        # print_time(self.name, self.delay, 5)
        if runThread(self.driver, self.timeout, self.xpath):
            if self.flag == "pinFlag":
                global pinFlag
                pinFlag = 1
            elif self.flag == "loginFlag":
                global loginFlag
                loginFlag = 1
            else:
                print("异常参数")
        print("退出线程：" + self.name)


def runThread(driver, timeout, xpath):
    try:
        WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, xpath)))
        return True
    except:
        return False


def enterPassword(password):
    print("请输入密码:")
    # password = []
    while 1:
        ch = msvcrt.getch()
        # 回车
        if ch == b'\r':
            msvcrt.putch(b'\n')
            # print('输入的密码是：%s' % b''.join(li).decode())
            return True
        # 退格
        elif ch == b'\x08':
            if password:
                password.pop()
                msvcrt.putch(b'\b')
                msvcrt.putch(b' ')
                msvcrt.putch(b'\b')
        # Esc
        elif ch == b'\x1b':
            return False
        else:
            password.append(ch)
            msvcrt.putch(b'*')


def loginPinterest(driver, account, password):
    global loginFlag
    threads = []
    driver.get("https://www.pinterest.com/")
    driver.find_element_by_xpath(
        '//*[@id="__PWS_ROOT__"]/div[1]/div/div/main/div[1]/div[1]/div[2]/div[2]/button/div/div').click()
    driver.find_element_by_xpath('//*[@id="email"]').send_keys(account)
    driver.find_element_by_xpath('//*[@id="password"]').send_keys(password)
    time.sleep(0.5)
    driver.find_element_by_xpath(
        '//*[@id="__PWS_ROOT__"]/div[1]/div/div/main/div[1]/div[2]/div[2]/div/div/div/div/div/div/div/div[4]/form/div[5]/button/div').click()

    possible_xpaths = ['//*[@id="HeaderContent"]/div/div/div[2]/div/div/div/div[5]/div[5]/button/div/div',
                       '//*[@id="HeaderContent"]/div/div/div/div/div[2]/div/div/div/div[10]/button/div/div']
    for i in range(len(possible_xpaths)):
        thread = detectThread(1, "Thread-" + str(i), possible_xpaths[i], driver, 60, "loginFlag")
        threads.append(thread)
    for t in threads:
        t.start()
    count_time = 0
    time_start = time.perf_counter()
    while count_time < 60:
        if loginFlag == 1:
            break
        count_time = time.perf_counter() - time_start
    print(count_time)
    if loginFlag == 1:
        loginFlag = 0
        return True
    else:
        loginFlag = 0
        print("登陆失败")
        return False


def amazon2pinterest(driver, category, ASIN):
    global pinFlag
    threads = []
    url_head = "https://www.amazon.com/Round-shell-pendant-necklace-choker/dp/"
    url_end = "/ref=sr_1_1?crid=2QCVNOFJ4EHEV&keywords=B09T35FRT7&qid=1647844910&sprefix=b09t35frt7%2Caps%2C527&sr=8-1"
    url_ASIN = ASIN
    url = url_head + url_ASIN + url_end
    driver.get(url)
    htmltext = driver.page_source
    if "We couldn't find that page" in htmltext:
        print("ASIN变狗")
        return -2
    try:
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="pinterest"]/i')))  # 元素是否可见
        driver.find_element_by_xpath('//*[@id="pinterest"]/i').click()
        handles = driver.window_handles
        driver.switch_to.window(handles[1])
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="pickerSearchField"]')))  # 元素是否可见
        driver.find_element_by_xpath('//*[@id="pickerSearchField"]').send_keys(category)

        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,
            '//*[@id="__PWS_ROOT__"]/div[1]/div[2]/div/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div/div/div')))  # 元素是否可见
        driver.find_element_by_xpath(
            '//*[@id="__PWS_ROOT__"]/div[1]/div[2]/div/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div/div/div').click()

        possible_xpaths = ['/html/body/div[3]/div/div/div/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/a/div/div/div/div/div[1]/img',
                           '/html/body/div[4]/div/div/div/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/a/div/div/div/div/div[1]/img'
                           '//*[@id="__PWS_ROOT__"]/div[1]/div[2]/div/div/div/div/div/div/div[2]/div/div[1]/div/div[1]/a/div/div/div/div/div[1]/img',
                           '/html/body/div[3]/div/div/div/div[2]/div/div/div/div[2]/div[4]/div[1]/a/div',
                           '/html/body/div[4]/div/div/div/div[2]/div/div/div/div[1]/img']
        for i in range(len(possible_xpaths)):
            thread = detectThread(1, "Thread-" + str(i), possible_xpaths[i], driver, 10, "pinFlag")
            threads.append(thread)
        for t in threads:
            t.start()
        count_time = 0
        time_start = time.perf_counter()
        while count_time < 10:
            if pinFlag == 1:
                break
            count_time = time.perf_counter() - time_start
        print(count_time)
        driver.close()
        driver.switch_to.window(handles[0])
        if pinFlag == 1:
            pinFlag = 0
            return 0
        else:
            pinFlag = 0
            return -1
    except Exception as e:
        print("未找到元素")
        driver.close()
        driver.switch_to.window(handles[0])
        pinFlag = 0
        return -1


if __name__ == '__main__':
    fileName = input("\n文件路径：\n")
    wb = load_workbook(fileName.replace("\"", ""))
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    print(ws.title)
    print(ws.max_row)
    print(ws.max_column)
    row = ws.max_row
    column = ws.max_column
    category = ws.cell(1, 1).value
    if ws.cell(1, 2).value is not None:
        account = ws.cell(1, 2).value
    else:
        account = input("输入账号:")
    if ws.cell(1, 3).value is not None:
        password = ws.cell(1, 3).value
    else:
        password = []
        enterPassword(password)
        password = b''.join(password).decode()
    driver = webdriver.Chrome()
    driver.maximize_window()
    loginPinterest(driver, account, password)
    try:
        for i in range(2, row + 1):
            print(i)
            result = amazon2pinterest(driver, category, ws.cell(i, 1).value)
            if result == 0:
                ws.cell(i, 2).value = '钉成功'
            elif result == -1:
                ws.cell(i, 2).value = '钉失败'
            elif result == -2:
                ws.cell(i, 2).value = 'ASIN变狗'
            else:
                ws.cell(i, 2).value = '未知原因'
            wb.save(fileName.replace("\"", ""))
        driver.quit()
    except:
        input("ERROR")



