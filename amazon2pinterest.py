from openpyxl import load_workbook
from selenium import webdriver
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


def loginPinterest(driver, account, password):
    driver.get("https://www.pinterest.com/")
    driver.find_element_by_xpath(
        '//*[@id="__PWS_ROOT__"]/div[1]/div/div/main/div[1]/div[1]/div[2]/div[2]/button/div/div').click()
    driver.find_element_by_xpath('//*[@id="email"]').send_keys(account)
    driver.find_element_by_xpath('//*[@id="password"]').send_keys(password)
    time.sleep(0.5)
    driver.find_element_by_xpath(
        '//*[@id="__PWS_ROOT__"]/div[1]/div/div/main/div[1]/div[2]/div[2]/div/div/div/div/div/div/div/div[4]/form/div[5]/button/div').click()
    try:
        WebDriverWait(driver, 30).until(EC.visibility_of_element_located(
            (By.XPATH, '//*[@id="HeaderContent"]/div/div/div[2]/div/div/div/div[5]/div[5]/button/div/div')))  # 元素是否可见
        return True
    except:
        print("登陆失败")
        return False


def amazon2pinterest(driver, category, ASIN):
    url_head = "https://www.amazon.com/Round-shell-pendant-necklace-choker/dp/"
    url_end = "/ref=sr_1_1?crid=2QCVNOFJ4EHEV&keywords=B09T35FRT7&qid=1647844910&sprefix=b09t35frt7%2Caps%2C527&sr=8-1"
    url_ASIN = ASIN
    url = url_head + url_ASIN + url_end
    driver.get(url)
    try:
        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="pinterest"]/i')))  # 元素是否可见
        driver.find_element_by_xpath('//*[@id="pinterest"]/i').click()
        handles = driver.window_handles
        driver.switch_to.window(handles[1])
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="pickerSearchField"]')))  # 元素是否可见
        driver.find_element_by_xpath('//*[@id="pickerSearchField"]').send_keys(category)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
            '//*[@id="__PWS_ROOT__"]/div[1]/div[2]/div/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div/div/div')))  # 元素是否可见
        driver.find_element_by_xpath(
            '//*[@id="__PWS_ROOT__"]/div[1]/div[2]/div/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div[2]/div/div/div/div').click()
        try:
            WebDriverWait(driver, 5).until(
                lambda driver: driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/a/div/div/div/div/div[1]/img')
                           or driver.find_element_by_xpath('/html/body/div[4]/div/div/div/div[2]/div/div/div/div[1]/img'))
        except:
            driver.close()
            driver.switch_to.window(handles[0])
            return True
        # WebDriverWait(driver, 10).until(
        #     EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div/div/div/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/a/div/div/div/div/div[1]/img')))  # 元素是否可见
        driver.close()
        driver.switch_to.window(handles[0])
        return True
    except Exception as e:
        print("未找到元素")
        driver.close()
        driver.switch_to.window(handles[0])
        return False


if __name__ == '__main__':
    fileName = input("\n文件路径：\n")
    wb = load_workbook(fileName.replace("\"", ""))
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    print(ws.title)
    print(ws.max_row)
    print(ws.max_column)
    row = ws.max_row
    column = ws.max_column
    category = ws.cell(1, 1).value
    if ws.cell(1, 2).value is not None:
        account = ws.cell(1, 2).value
    else:
        account = input("输入账号:")
    if ws.cell(1, 3).value is not None:
        password = ws.cell(1, 3).value
    else:
        password = input("输入密码:")
    driver = webdriver.Chrome()
    driver.maximize_window()
    loginPinterest(driver, account, password)
    try:
        for i in range(2, row + 1):
            print(i)
            result = amazon2pinterest(driver, category, ws.cell(i, 1).value)
            if not result:
                ws.cell(i, 2).value = '钉失败'
            else:
                ws.cell(i, 2).value = '钉成功'
            wb.save(fileName.replace("\"", ""))
        driver.quit()
    except:
        input("ERROR")



